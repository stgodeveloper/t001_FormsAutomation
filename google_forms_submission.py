import logging
import time
import os
from openpyxl import load_workbook
import requests
import tkinter as tk
from tkinter import messagebox

logging.info("---- Starting module 'google_forms_submission' ----")

# Configurable parameters
PROCESS_DATA_DIRECTORY = "process_data"
WORKTRAY_FILE = "worktray.xlsx"
GOOGLE_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSf_5o0pOYiDzAJp2uRdSfoj5xxIfzFs0M9beiaXTsdFgeAcrw/formResponse"  # Form submission URL
SUBMISSION_DELAY = 1  # Delay between submissions (in seconds)
LOGS_DIRECTORY = "_logs"  # Directory to store logs
LOGS_FILE = os.path.join(LOGS_DIRECTORY, "google_forms_submission.log")  # Log file path

# Status messages
SUCCESS_MESSAGE = "Ingreso exitoso a Forms"
FAILURE_MESSAGE = "Error en el ingreso a Forms"
NETWORK_ERROR_MESSAGE = "Error de conexión (revise su conexión a internet o la URL del formulario)"
BROWSER_ERROR_MESSAGE = "Error de navegador (no se pudo acceder al formulario)"

def configure_logging():
    """
    Configures logging for the application.
    """
    if not os.path.exists(LOGS_DIRECTORY):
        os.makedirs(LOGS_DIRECTORY)
    
    # Clear any existing logging configuration
    logging.basicConfig(
        filename=LOGS_FILE,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        force=True  # Force reconfiguration of logging
    )
    logging.info("Logging configured successfully.")

def show_results_popup(success_count, failure_count):
    """
    Shows a popup with the results of the execution.
    """
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Create a Toplevel window for the popup
    popup = tk.Toplevel(root)
    popup.title("Resultados de la Ejecución")
    popup.geometry("400x150")  # Set the size of the popup

    # Add a label with the message
    message = f"Resultados de la ejecución:\n\n- Filas cargadas exitosamente: {success_count}\n- Filas no cargadas: {failure_count}"
    label = tk.Label(popup, text=message)
    label.pack(pady=10)

    # Function to handle the "Ok" button
    def on_ok():
        popup.destroy()
        root.quit()

    # Add an "Ok" button
    ok_button = tk.Button(popup, text="Ok", command=on_ok)
    ok_button.pack(pady=10)

    # Center the popup
    popup.update_idletasks()
    width = popup.winfo_width()
    height = popup.winfo_height()
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    popup.geometry(f"+{x}+{y}")

    # Wait for the user to interact with the popup
    popup.mainloop()

def submit_to_google_forms():
    """
    Submits data from the worktray to Google Forms.
    Updates "Ingreso exitoso a Forms" and "Observaciones" based on the submission result.
    Skips rows where "Ingreso exitoso a Forms" is already TRUE.
    Handles network errors, browser errors, and other exceptions.
    Only processes rows where "Datos correctos" is TRUE.
    """
    try:
        # Path to the worktray file
        worktray_path = os.path.join(PROCESS_DATA_DIRECTORY, WORKTRAY_FILE)
        
        # Read the worktray
        logging.info(f"Reading the worktray from: {worktray_path}")
        worktray_wb = load_workbook(worktray_path)
        worktray_ws = worktray_wb.active
        
        # Log the number of rows to process
        total_rows = worktray_ws.max_row - 1  # Subtract header row
        logging.info(f"Total rows to process: {total_rows}")

        # Counters for successful and failed submissions
        success_count = 0
        failure_count = 0
        
        # Iterate through rows (skip the header row)
        for row in worktray_ws.iter_rows(min_row=2, max_row=worktray_ws.max_row, min_col=1, max_col=9):
            row_number = row[0].row - 1  # Adjust for zero-based index
            logging.info(f"Processing row {row_number}: {[cell.value for cell in row]}")
            
            # Skip rows where "Datos correctos" is FALSE
            if row[4].value != True:
                logging.info(f"Skipping row {row_number}: 'Datos correctos' is FALSE")
                failure_count += 1
                continue
            
            # Skip rows where "Ingreso exitoso a Forms" is already TRUE
            if row[5].value == True:
                logging.info(f"Skipping row {row_number}: 'Ingreso exitoso a Forms' is already TRUE")
                success_count += 1
                continue
            
            try:
                # Prepare data for Google Forms submission
                form_data = {
                    "entry.274949855": row[0].value,  # Nombre
                    "entry.1623880646": row[1].value,  # Producto
                    "entry.1721353382": row[2].value,  # Monto
                    "entry.1896335859": row[3].value,  # Fecha de Solicitud
                }
                
                logging.info(f"Submitting row {row_number} to Google Forms: {form_data}")
                
                # Submit data to Google Forms
                response = requests.post(
                    GOOGLE_FORM_URL,
                    data=form_data,
                    headers={"Content-Type": "application/x-www-form-urlencoded"},
                    timeout=10  # Set a timeout for the request
                )
                
                # Check if the submission was successful
                if response.status_code == 200 or "Gracias" in response.text:  # Google Forms may return a 200 or a redirect
                    row[5].value = True  # Ingreso exitoso a Forms
                    row[6].value = SUCCESS_MESSAGE  # Observaciones
                    logging.info(f"Row {row_number} submitted successfully. Response: {response.status_code}")
                    success_count += 1
                else:
                    row[5].value = False  # Ingreso exitoso a Forms
                    row[6].value = FAILURE_MESSAGE  # Observaciones
                    logging.error(f"Error submitting row {row_number}. Status code: {response.status_code}, Response text: {response.text}")
                    failure_count += 1
            
            except requests.exceptions.Timeout:
                # Handle timeout errors (e.g., no internet connection)
                row[5].value = False  # Ingreso exitoso a Forms
                row[6].value = NETWORK_ERROR_MESSAGE  # Observaciones
                logging.error(f"Timeout error submitting row {row_number}: No internet connection or server took too long to respond.")
                failure_count += 1
            
            except requests.exceptions.ConnectionError:
                # Handle connection errors (e.g., invalid URL or no internet)
                row[5].value = False  # Ingreso exitoso a Forms
                row[6].value = NETWORK_ERROR_MESSAGE  # Observaciones
                logging.error(f"Connection error submitting row {row_number}: Invalid URL or no internet connection.")
                failure_count += 1
            
            except requests.exceptions.RequestException as e:
                # Handle other request-related errors (e.g., browser errors)
                row[5].value = False  # Ingreso exitoso a Forms
                row[6].value = BROWSER_ERROR_MESSAGE  # Observaciones
                logging.error(f"Browser error submitting row {row_number}: {str(e)}", exc_info=True)
                failure_count += 1
            
            except Exception as e:
                # Handle any other unexpected errors
                row[5].value = False  # Ingreso exitoso a Forms
                row[6].value = FAILURE_MESSAGE  # Observaciones
                logging.error(f"Unexpected error submitting row {row_number}: {str(e)}", exc_info=True)
                failure_count += 1
            
            # Add delay between submissions
            time.sleep(SUBMISSION_DELAY)
        
        # Save the updated worktray
        worktray_wb.save(worktray_path)
        logging.info(f"Worktray updated and saved to: {worktray_path}")

        # Show the results popup
        show_results_popup(success_count, failure_count)
        
        return True
    
    except Exception as e:
        logging.error(f"Error during Google Forms submission: {str(e)}", exc_info=True)
        return False

# Execute the function if the script is run directly
if __name__ == "__main__":
    # Configure logging
    if not os.path.exists("_logs"):
        os.makedirs("_logs")
    configure_logging()
    logging.info("---- Starting module 'google_forms_submission' ----")
    # Submit data to Google Forms
    success = submit_to_google_forms()
    
    if not success:
        logging.error("Google Forms submission failed. Check the logs for details.")
    else:
        logging.info("Google Forms submission completed successfully.")