import logging
import os
from openpyxl import load_workbook
from datetime import datetime
import re  # Regular expressions for special character validation

logging.info("---- Starting module 'input_file_processment' ----")

# Configurable parameters
PROCESS_DATA_DIRECTORY = "process_data"
WORKTRAY_FILE = "worktray.xlsx"

# Validation messages
INCOMPLETE_DATA_MESSAGE = "Faltan datos" 
INVALID_DATE_MESSAGE = "Fecha de Solicitud no está en formato fecha"
INVALID_NAME_TEXT_MESSAGE = "Ingrese Nombre válido"
INVALID_PRODUCT_TEXT_MESSAGE = "Ingrese un Producto válido"
INVALID_NUMBER_MESSAGE = "Ingrese un monto válido"
INVALID_SPECIAL_CHARACTERS_MESSAGE = "Caracteres especiales no permitidos"

def is_excel_date(value):
    """
    Check if the value is a valid Excel date (numeric format).
    Excel stores dates as numbers, where 1 = 1900-01-01.
    """
    if isinstance(value, (int, float)):
        try:
            # Convert Excel numeric date to datetime
            datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(value) - 2)
            return True
        except ValueError:
            return False
    return False

def is_datetime_object(value):
    """
    Check if the value is already a datetime object.
    """
    return isinstance(value, datetime)

def is_date_string(value):
    """
    Check if the value is a string that represents a valid date in the format "dd-mm-yyyy".
    """
    if isinstance(value, str):
        try:
            datetime.strptime(value, "%d-%m-%Y")
            return True
        except ValueError:
            return False
    return False

def is_text(value):
    """
    Check if the value is a valid text (string).
    """
    return isinstance(value, str)

def is_number(value):
    """
    Check if the value is a valid number (integer or float).
    """
    return isinstance(value, (int, float))

def contains_special_characters(value):
    """
    Check if the value contains special characters that are not allowed.
    Only letters, spaces, and some special characters (like accents) are allowed.
    """
    if not isinstance(value, str):
        return False
    # Regular expression to allow letters, spaces, and some special characters (like accents)
    pattern = r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s\-\.]+$'
    return not bool(re.match(pattern, value))

def validate_worktray():
    """
    Validates all rows in the worktray. If any field is empty, sets "Datos correctos" to FALSE
    and adds a comment in the "Observaciones" column. Also validates the date format in "Fecha de Solicitud",
    text format in "Nombre" and "Producto", number format in "Monto", and special characters in "Nombre" and "Producto".
    Maintains the original formatting.
    """
    try:
        # Path to the worktray file
        worktray_path = os.path.join(PROCESS_DATA_DIRECTORY, WORKTRAY_FILE)
        
        # Read the worktray
        logging.info(f"Reading the worktray from: {worktray_path}")
        worktray_wb = load_workbook(worktray_path)
        worktray_ws = worktray_wb.active
        
        # Iterate through rows (skip the header row)
        for row in worktray_ws.iter_rows(min_row=2, max_row=worktray_ws.max_row, min_col=1, max_col=7):
            is_valid = True
            observations = []
            
            # Check for empty fields (columns A to D: Nombre, Producto, Monto, Fecha de Solicitud)
            for cell in row[:4]:  # Columns A to D
                if cell.value is None or str(cell.value).strip() == "":
                    is_valid = False
                    observations.append(INCOMPLETE_DATA_MESSAGE)
                    break  # Stop checking other fields if one is empty

            # Validate "Nombre" (column A) and "Producto" (column B) as text
            nombre = row[0].value  # Column A
            producto = row[1].value  # Column B
            if nombre and not is_text(nombre):
                is_valid = False
                observations.append(f"{INVALID_NAME_TEXT_MESSAGE}")
            elif nombre and contains_special_characters(nombre):
                is_valid = False
                observations.append(f"{INVALID_SPECIAL_CHARACTERS_MESSAGE}")
            if producto is None or str(producto).strip() == "":
                is_valid = False
                observations.append(f"{INVALID_PRODUCT_TEXT_MESSAGE}")
            
            # Validate "Monto" (column C) as a number
            monto = row[2].value  # Column C
            if monto and not is_number(monto):
                is_valid = False
                observations.append(f"{INVALID_NUMBER_MESSAGE}")
            
            # Validate date format in "Fecha de Solicitud" (column D)
            fecha_solicitud = row[3].value
            if fecha_solicitud:
                # Check if the value is a valid Excel date, datetime object, or a valid date string
                if not (is_excel_date(fecha_solicitud) or is_datetime_object(fecha_solicitud)):
                    if is_date_string(fecha_solicitud):
                        # If it's a date string, it's invalid because we expect a real date, not a string
                        is_valid = False
                        observations.append(INVALID_DATE_MESSAGE)
                    else:
                        # If it's not a date string, it's also invalid
                        is_valid = False
                        observations.append(INVALID_DATE_MESSAGE)
            else:
                # If fecha_solicitud is empty, it's already handled by the empty field check
                pass
            
            # Update "Datos correctos" (column E) and "Observaciones" (column G)
            row[4].value = is_valid  # Datos correctos (column E)
            row[6].value = "; ".join(observations) if observations else ""  # Observaciones (column G)
        
        # Save the updated worktray
        worktray_wb.save(worktray_path)
        logging.info(f"Worktray validation completed and saved to: {worktray_path}")
        
        return True
    
    except Exception as e:
        logging.error(f"Error validating the worktray: {str(e)}")
        return False

# Execute the function if the script is run directly
if __name__ == "__main__":
    # Configure logging
    if not os.path.exists("_logs"):
        os.makedirs("_logs")

    logging.basicConfig(
        filename="_logs/input_file_processment.log",
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )
    
    # Validate the worktray
    success = validate_worktray()
    
    if not success:
        logging.error("Worktray validation failed. Check the logs for details.")
    else:
        logging.info("Worktray validation completed successfully.")